from __future__ import annotations

import os
import uuid
from datetime import UTC, date, datetime
from decimal import Decimal
from typing import Any

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.encryption import decrypt
from app.core.visibility import VisibilityContext
from app.db.models.account import Account
from app.db.models.category import Category
from app.db.models.debt import Debt
from app.db.models.export_job import ExportJob
from app.db.models.fire import FireScenario
from app.db.models.snapshot import AccountSnapshot
from app.db.models.transaction import Transaction
from app.repositories.account import AccountRepository

LIGHT_GRAY = "F2F2F2"
HEADER_BLUE = "1E3A5F"


def _cell_money(ws: Any, row: int, col: int, value: Decimal) -> None:
    cell = ws.cell(row=row, column=col, value=float(value))
    cell.number_format = "$#,##0.00"


def _cell_date(ws: Any, row: int, col: int, value: date) -> None:
    cell = ws.cell(row=row, column=col, value=value)
    cell.number_format = "YYYY-MM-DD"


def _bold_header_row(ws: Any, headers: list[str]) -> None:
    from openpyxl.styles import Font, PatternFill  # type: ignore[import-untyped]

    fill = PatternFill(start_color=HEADER_BLUE, end_color=HEADER_BLUE, fill_type="solid")
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill


def _alternate_fill(ws: Any, row: int, num_cols: int) -> None:
    from openpyxl.styles import PatternFill  # type: ignore[import-untyped,unused-ignore]

    if row % 2 == 0:
        fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
        for col in range(1, num_cols + 1):
            ws.cell(row=row, column=col).fill = fill


async def _fetch_accounts(session: AsyncSession, ctx: VisibilityContext) -> list[Account]:
    repo = AccountRepository(session)
    return await repo.get_visible(ctx, is_active=True)


async def _latest_balance(
    session: AsyncSession, account_id: uuid.UUID, as_of: date
) -> Decimal | None:
    result = await session.execute(
        select(AccountSnapshot.balance)
        .where(
            AccountSnapshot.account_id == account_id,
            AccountSnapshot.snapshot_date <= as_of,
        )
        .order_by(AccountSnapshot.snapshot_date.desc())
        .limit(1)
    )
    return result.scalar_one_or_none()


async def _fetch_net_worth_series(
    session: AsyncSession,
    accounts: list[Account],
    from_date: date,
    to_date: date,
) -> list[tuple[date, Decimal, Decimal]]:
    """Returns list of (month_end, total_assets, total_liabilities)."""
    import calendar

    liability_types = {
        "mortgage",
        "credit_card",
        "auto_loan",
        "personal_loan",
        "student_loan",
        "other_liability",
        "heloc",
    }
    series: list[tuple[date, Decimal, Decimal]] = []
    y, m = from_date.year, from_date.month
    while date(y, m, 1) <= to_date:
        month_end = date(y, m, calendar.monthrange(y, m)[1])
        total_assets = Decimal("0")
        total_liabs = Decimal("0")
        for acct in accounts:
            if not acct.include_in_net_worth:
                continue
            bal = await _latest_balance(session, acct.id, month_end)
            if bal is None:
                continue
            if acct.account_type in liability_types:
                total_liabs += abs(bal)
            else:
                total_assets += bal
        series.append((month_end, total_assets, total_liabs))
        y, m = (y + 1, 1) if m == 12 else (y, m + 1)
    return series


async def _fetch_transactions(
    session: AsyncSession,
    account_ids: list[uuid.UUID],
    from_date: date,
    to_date: date,
) -> list[Transaction]:
    if not account_ids:
        return []
    result = await session.execute(
        select(Transaction)
        .where(
            Transaction.account_id.in_(account_ids),
            Transaction.transaction_date >= from_date,
            Transaction.transaction_date <= to_date,
        )
        .order_by(Transaction.transaction_date.desc())
        .limit(50000)
    )
    return list(result.scalars().all())


async def _fetch_categories(
    session: AsyncSession, household_id: uuid.UUID
) -> dict[uuid.UUID, Category]:
    result = await session.execute(select(Category).where(Category.household_id == household_id))
    return {c.id: c for c in result.scalars().all()}


async def _fetch_debts(session: AsyncSession, account_ids: list[uuid.UUID]) -> list[Debt]:
    if not account_ids:
        return []
    result = await session.execute(select(Debt).where(Debt.account_id.in_(account_ids)))
    return list(result.scalars().all())


async def _fetch_fire_scenarios(
    session: AsyncSession, household_id: uuid.UUID
) -> list[FireScenario]:
    result = await session.execute(
        select(FireScenario).where(FireScenario.household_id == household_id)
    )
    return list(result.scalars().all())


def _build_net_worth_sheet(wb: Any, series: list[tuple[date, Decimal, Decimal]]) -> None:
    ws = wb.create_sheet("Net Worth History")
    headers = ["Month End", "Total Assets", "Total Liabilities", "Net Worth"]
    _bold_header_row(ws, headers)
    for row_idx, (month_end, assets, liabs) in enumerate(series, start=2):
        _cell_date(ws, row_idx, 1, month_end)
        _cell_money(ws, row_idx, 2, assets)
        _cell_money(ws, row_idx, 3, liabs)
        _cell_money(ws, row_idx, 4, assets - liabs)
        _alternate_fill(ws, row_idx, 4)


def _build_accounts_sheet(
    wb: Any,
    accounts: list[Account],
    balances: dict[uuid.UUID, Decimal],
    anonymized: bool,
) -> None:
    ws = wb.create_sheet("Account Directory")
    headers = ["Nickname", "Type", "Institution", "Account Number", "Balance", "Active"]
    _bold_header_row(ws, headers)
    for row_idx, acct in enumerate(accounts, start=2):
        inst = "N/A"
        if acct.institution_name_enc:
            try:
                inst = decrypt(acct.institution_name_enc)
            except Exception:
                inst = "N/A"

        acct_num: str
        if acct.account_number_enc:
            try:
                full = decrypt(acct.account_number_enc)
                acct_num = f"••••{full[-4:]}" if anonymized and len(full) >= 4 else full
            except Exception:
                acct_num = "N/A"
        else:
            acct_num = "N/A"

        bal = balances.get(acct.id, Decimal("0"))
        ws.cell(row=row_idx, column=1, value=acct.nickname)
        ws.cell(row=row_idx, column=2, value=acct.account_type)
        ws.cell(row=row_idx, column=3, value=inst)
        ws.cell(row=row_idx, column=4, value=acct_num)
        _cell_money(ws, row_idx, 5, bal)
        ws.cell(row=row_idx, column=6, value="Yes" if acct.is_active else "No")
        _alternate_fill(ws, row_idx, 6)


def _build_transactions_sheet(
    wb: Any,
    transactions: list[Transaction],
    accounts: list[Account],
    cat_map: dict[uuid.UUID, Category],
) -> None:
    ws = wb.create_sheet("Transactions")
    headers = ["Date", "Account", "Payee", "Category", "Amount", "Transfer", "Tags"]
    _bold_header_row(ws, headers)
    acct_map = {a.id: a for a in accounts}
    for row_idx, txn in enumerate(transactions, start=2):
        acct = acct_map.get(txn.account_id)
        acct_name = acct.nickname if acct else str(txn.account_id)
        cat = cat_map.get(txn.category_id) if txn.category_id else None
        cat_name = cat.name if cat else ""
        _cell_date(ws, row_idx, 1, txn.transaction_date)
        ws.cell(row=row_idx, column=2, value=acct_name)
        ws.cell(row=row_idx, column=3, value=txn.payee_normalized or txn.payee_raw or "")
        ws.cell(row=row_idx, column=4, value=cat_name)
        _cell_money(ws, row_idx, 5, txn.amount)
        ws.cell(row=row_idx, column=6, value="Yes" if txn.is_transfer else "No")
        ws.cell(row=row_idx, column=7, value=", ".join(txn.tags) if txn.tags else "")
        _alternate_fill(ws, row_idx, 7)

    if len(transactions) > 0:
        ws.auto_filter.ref = ws.dimensions


def _build_budget_vs_actuals_sheet(
    wb: Any,
    transactions: list[Transaction],
    cat_map: dict[uuid.UUID, Category],
) -> None:
    """Builds a monthly budget vs actuals summary sheet."""
    from collections import defaultdict

    ws = wb.create_sheet("Budget vs Actuals")
    headers = ["Category", "Total Spending"]
    _bold_header_row(ws, headers)

    sums: dict[str, Decimal] = defaultdict(lambda: Decimal("0"))
    for txn in transactions:
        if txn.is_transfer or txn.amount >= 0:
            continue
        cat = cat_map.get(txn.category_id) if txn.category_id else None
        name = cat.name if cat else "Uncategorized"
        sums[name] += abs(txn.amount)

    for row_idx, (name, total) in enumerate(
        sorted(sums.items(), key=lambda kv: kv[1], reverse=True), start=2
    ):
        ws.cell(row=row_idx, column=1, value=name)
        _cell_money(ws, row_idx, 2, total)
        _alternate_fill(ws, row_idx, 2)


def _build_spending_sheet(
    wb: Any,
    transactions: list[Transaction],
    cat_map: dict[uuid.UUID, Category],
) -> None:
    from collections import defaultdict

    ws = wb.create_sheet("Spending by Category")
    headers = ["Category", "Amount", "Transaction Count"]
    _bold_header_row(ws, headers)

    sums: dict[str, list[Any]] = defaultdict(lambda: [Decimal("0"), 0])
    for txn in transactions:
        if txn.is_transfer or txn.amount >= 0:
            continue
        cat = cat_map.get(txn.category_id) if txn.category_id else None
        name = cat.name if cat else "Uncategorized"
        sums[name][0] += abs(txn.amount)
        sums[name][1] += 1

    for row_idx, (name, (total, count)) in enumerate(
        sorted(sums.items(), key=lambda kv: kv[1][0], reverse=True), start=2
    ):
        ws.cell(row=row_idx, column=1, value=name)
        _cell_money(ws, row_idx, 2, total)
        ws.cell(row=row_idx, column=3, value=count)
        _alternate_fill(ws, row_idx, 3)


def _build_debt_sheet(wb: Any, debts: list[Debt], acct_map: dict[uuid.UUID, Account]) -> None:
    ws = wb.create_sheet("Debt Schedule")
    headers = ["Account", "Current Balance", "Interest Rate", "Minimum Payment"]
    _bold_header_row(ws, headers)
    for row_idx, debt in enumerate(debts, start=2):
        acct = acct_map.get(debt.account_id)
        nick = acct.nickname if acct else str(debt.account_id)
        rate = float(debt.interest_rate * 100) if debt.interest_rate else 0.0
        minp = float(debt.minimum_payment) if debt.minimum_payment else 0.0
        ws.cell(row=row_idx, column=1, value=nick)
        _cell_money(ws, row_idx, 2, debt.current_balance)
        ws.cell(row=row_idx, column=3, value=f"{rate:.2f}%")
        _cell_money(ws, row_idx, 4, Decimal(str(minp)))
        _alternate_fill(ws, row_idx, 4)


def _build_fire_sheet(wb: Any, scenarios: list[FireScenario]) -> None:
    ws = wb.create_sheet("FIRE Projections")
    headers = [
        "Scenario",
        "Target Annual Spend",
        "Safe Withdrawal Rate",
        "Expected Return",
        "Inflation Rate",
        "Target Retirement Age",
    ]
    _bold_header_row(ws, headers)
    for row_idx, sc in enumerate(scenarios, start=2):
        ws.cell(row=row_idx, column=1, value=sc.name)
        _cell_money(ws, row_idx, 2, sc.target_annual_spend)
        ws.cell(row=row_idx, column=3, value=f"{float(sc.safe_withdrawal_rate) * 100:.1f}%")
        ws.cell(row=row_idx, column=4, value=f"{float(sc.expected_annual_return) * 100:.1f}%")
        ws.cell(row=row_idx, column=5, value=f"{float(sc.expected_inflation_rate) * 100:.1f}%")
        ws.cell(row=row_idx, column=6, value=sc.target_retirement_age)
        _alternate_fill(ws, row_idx, 6)


async def generate(job: ExportJob, session: AsyncSession, output_dir: str) -> str:
    """Generate an Excel export file. Returns the filename (not full path)."""
    from_date = date.fromisoformat(job.parameters["from_date"])
    to_date = date.fromisoformat(job.parameters["to_date"])
    anonymized = job.anonymized
    role = job.parameters.get("role", "primary")
    member_id_str = job.parameters.get("member_id")
    member_id = uuid.UUID(member_id_str) if member_id_str else None

    ctx = VisibilityContext(
        user_id=job.generated_by,
        member_id=member_id,
        role=role,
        household_id=job.household_id,
    )

    generated_at = datetime.now(UTC)
    ts = generated_at.strftime("%Y-%m-%dT%H-%M-%SZ")
    suffix = "summary" if anonymized else "executor"
    filename = f"hearthledger_excel_{suffix}_{ts}.xlsx"
    output_path = os.path.join(output_dir, filename)
    os.makedirs(output_dir, exist_ok=True)

    # Fetch data
    accounts = await _fetch_accounts(session, ctx)
    account_ids = [a.id for a in accounts]
    cat_map = await _fetch_categories(session, job.household_id)
    today = date.today()

    account_balances: dict[uuid.UUID, Decimal] = {}
    for acct in accounts:
        bal = await _latest_balance(session, acct.id, today)
        account_balances[acct.id] = bal if bal is not None else Decimal("0")

    nw_series = await _fetch_net_worth_series(session, accounts, from_date, to_date)
    transactions = await _fetch_transactions(session, account_ids, from_date, to_date)
    debts = await _fetch_debts(session, account_ids)
    scenarios = await _fetch_fire_scenarios(session, job.household_id)

    from openpyxl import Workbook  # type: ignore[import-untyped]

    wb: Any = Workbook()
    # Remove default sheet
    if wb.active is not None:
        del wb[wb.active.title]

    _build_net_worth_sheet(wb, nw_series)
    _build_accounts_sheet(wb, accounts, account_balances, anonymized)
    _build_transactions_sheet(wb, transactions, accounts, cat_map)
    _build_budget_vs_actuals_sheet(wb, transactions, cat_map)
    _build_spending_sheet(wb, transactions, cat_map)

    acct_map = {a.id: a for a in accounts}
    _build_debt_sheet(wb, debts, acct_map)

    if scenarios:
        _build_fire_sheet(wb, scenarios)

    wb.save(output_path)
    return filename
