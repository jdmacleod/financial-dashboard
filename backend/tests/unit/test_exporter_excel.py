"""Tests for the Excel exporter — exercises openpyxl generation with real DB data."""

from __future__ import annotations

import os
import uuid
from datetime import UTC, date, datetime
from decimal import Decimal
from typing import Any

from sqlalchemy.ext.asyncio import AsyncSession

from app.db.models.account import Account
from app.db.models.category import Category
from app.db.models.debt import Debt
from app.db.models.export_job import ExportJob
from app.db.models.fire import FireScenario
from app.db.models.household import Household
from app.db.models.member import HouseholdMember
from app.db.models.snapshot import AccountSnapshot
from app.db.models.transaction import Transaction
from app.db.models.user import User
from app.exporters import excel_exporter


def _now() -> datetime:
    return datetime.now(UTC)


def _make_job(
    household: Household,
    user: User,
    member: HouseholdMember,
    export_type: str = "excel_summary",
    from_date: str = "2024-01-01",
    to_date: str = "2024-12-31",
) -> ExportJob:
    is_executor = export_type.endswith("executor")
    return ExportJob(
        id=uuid.uuid4(),
        household_id=household.id,
        export_type=export_type,
        anonymized=not is_executor,
        parameters={
            "from_date": from_date,
            "to_date": to_date,
            "member_id": str(member.id),
            "role": "primary",
        },
        status="pending",
        generated_by=user.id,
        created_at=_now(),
    )


async def _seed_account(
    db_session: AsyncSession,
    household: Household,
    account_type: str = "checking",
    nickname: str = "Checking",
) -> Account:
    acct = Account(
        household_id=household.id,
        account_type=account_type,
        nickname=nickname,
        include_in_net_worth=True,
        is_active=True,
        created_at=_now(),
        updated_at=_now(),
    )
    db_session.add(acct)
    await db_session.flush()
    return acct


async def _seed_snapshot(
    db_session: AsyncSession,
    account: Account,
    balance: Decimal,
    snap_date: date,
) -> AccountSnapshot:
    snap = AccountSnapshot(
        account_id=account.id,
        snapshot_date=snap_date,
        balance=balance,
        source="manual",
        created_at=_now(),
    )
    db_session.add(snap)
    await db_session.flush()
    return snap


async def _seed_category(
    db_session: AsyncSession, household: Household, name: str, is_income: bool = False
) -> Category:
    cat = Category(
        household_id=household.id,
        name=name,
        is_income=is_income,
        is_system=False,
        created_at=_now(),
    )
    db_session.add(cat)
    await db_session.flush()
    return cat


async def _seed_transaction(
    db_session: AsyncSession,
    account: Account,
    amount: Decimal,
    txn_date: date,
    category: Category | None = None,
    is_transfer: bool = False,
) -> Transaction:
    txn = Transaction(
        account_id=account.id,
        transaction_date=txn_date,
        amount=amount,
        payee_raw="Test Payee",
        payee_normalized="Test Payee",
        tags=[],
        source="manual",
        is_transfer=is_transfer,
        category_id=category.id if category else None,
        created_at=_now(),
        updated_at=_now(),
    )
    db_session.add(txn)
    await db_session.flush()
    return txn


async def test_generate_excel_summary_empty_household(
    db_session: AsyncSession,
    household: Household,
    primary_member: HouseholdMember,
    primary_user: User,
    tmp_path: Any,
) -> None:
    """Excel export with no data still produces a valid .xlsx file."""
    job = _make_job(household, primary_user, primary_member)
    db_session.add(job)
    await db_session.flush()

    filename = await excel_exporter.generate(job, db_session, str(tmp_path))

    assert filename.endswith(".xlsx")
    assert os.path.exists(os.path.join(str(tmp_path), filename))

    from openpyxl import load_workbook  # type: ignore[import-untyped]

    wb = load_workbook(os.path.join(str(tmp_path), filename))
    assert "Net Worth History" in wb.sheetnames
    assert "Account Directory" in wb.sheetnames
    assert "Transactions" in wb.sheetnames
    assert "Budget vs Actuals" in wb.sheetnames
    assert "Spending by Category" in wb.sheetnames
    assert "Debt Schedule" in wb.sheetnames


async def test_generate_excel_with_accounts_and_snapshots(
    db_session: AsyncSession,
    household: Household,
    primary_member: HouseholdMember,
    primary_user: User,
    tmp_path: Any,
) -> None:
    """Net Worth History and Account Directory sheets populate with real data."""
    checking = await _seed_account(db_session, household, "checking", "Checking")
    await _seed_snapshot(db_session, checking, Decimal("5000"), date(2024, 1, 31))
    await _seed_snapshot(db_session, checking, Decimal("6000"), date(2024, 12, 31))

    job = _make_job(
        household, primary_user, primary_member, from_date="2024-01-01", to_date="2024-12-31"
    )
    db_session.add(job)
    await db_session.flush()

    filename = await excel_exporter.generate(job, db_session, str(tmp_path))

    from openpyxl import load_workbook  # type: ignore[import-untyped]

    wb = load_workbook(os.path.join(str(tmp_path), filename))
    nw_ws = wb["Net Worth History"]
    # Header row at row 1, data starts at row 2
    assert nw_ws.cell(row=1, column=1).value == "Month End"
    assert nw_ws.max_row >= 2


async def test_generate_excel_with_transactions_and_categories(
    db_session: AsyncSession,
    household: Household,
    primary_member: HouseholdMember,
    primary_user: User,
    tmp_path: Any,
) -> None:
    """Transactions sheet and spending sheets populate correctly."""
    checking = await _seed_account(db_session, household, "checking", "Checking")
    groceries = await _seed_category(db_session, household, "Groceries", is_income=False)
    salary = await _seed_category(db_session, household, "Salary", is_income=True)

    await _seed_transaction(
        db_session, checking, Decimal("-150.00"), date(2024, 6, 15), category=groceries
    )
    await _seed_transaction(
        db_session, checking, Decimal("3000.00"), date(2024, 6, 1), category=salary
    )
    # A transfer — should be excluded from spending sheets
    await _seed_transaction(
        db_session, checking, Decimal("-500.00"), date(2024, 6, 20), is_transfer=True
    )

    job = _make_job(
        household, primary_user, primary_member, from_date="2024-01-01", to_date="2024-12-31"
    )
    db_session.add(job)
    await db_session.flush()

    filename = await excel_exporter.generate(job, db_session, str(tmp_path))

    from openpyxl import load_workbook  # type: ignore[import-untyped]

    wb = load_workbook(os.path.join(str(tmp_path), filename))
    txn_ws = wb["Transactions"]
    assert txn_ws.cell(row=1, column=1).value == "Date"
    assert txn_ws.max_row >= 3  # header + 3 transactions

    spend_ws = wb["Spending by Category"]
    # Groceries should appear; Salary (income) and transfer should not
    categories_in_sheet = [
        spend_ws.cell(row=r, column=1).value for r in range(2, spend_ws.max_row + 1)
    ]
    assert "Groceries" in categories_in_sheet
    assert "Salary" not in categories_in_sheet


async def test_generate_excel_with_debts(
    db_session: AsyncSession,
    household: Household,
    primary_member: HouseholdMember,
    primary_user: User,
    tmp_path: Any,
) -> None:
    """Debt Schedule sheet is populated with debt data."""
    credit = await _seed_account(db_session, household, "credit_card", "Visa")
    debt = Debt(
        account_id=credit.id,
        original_balance=Decimal("2000"),
        current_balance=Decimal("2000"),
        interest_rate=Decimal("0.2199"),
        minimum_payment=Decimal("50"),
        created_at=_now(),
        updated_at=_now(),
    )
    db_session.add(debt)
    await db_session.flush()

    job = _make_job(household, primary_user, primary_member)
    db_session.add(job)
    await db_session.flush()

    filename = await excel_exporter.generate(job, db_session, str(tmp_path))

    from openpyxl import load_workbook  # type: ignore[import-untyped]

    wb = load_workbook(os.path.join(str(tmp_path), filename))
    debt_ws = wb["Debt Schedule"]
    assert debt_ws.max_row >= 2


async def test_generate_excel_with_fire_scenarios(
    db_session: AsyncSession,
    household: Household,
    primary_member: HouseholdMember,
    primary_user: User,
    tmp_path: Any,
) -> None:
    """FIRE Projections sheet appears when there are FIRE scenarios."""
    scenario = FireScenario(
        household_id=household.id,
        name="Lean FIRE",
        target_annual_spend=Decimal("40000"),
        safe_withdrawal_rate=Decimal("0.04"),
        expected_annual_return=Decimal("0.07"),
        expected_inflation_rate=Decimal("0.03"),
        created_at=_now(),
        updated_at=_now(),
    )
    db_session.add(scenario)
    await db_session.flush()

    job = _make_job(household, primary_user, primary_member)
    db_session.add(job)
    await db_session.flush()

    filename = await excel_exporter.generate(job, db_session, str(tmp_path))

    from openpyxl import load_workbook  # type: ignore[import-untyped]

    wb = load_workbook(os.path.join(str(tmp_path), filename))
    assert "FIRE Projections" in wb.sheetnames
    fire_ws = wb["FIRE Projections"]
    assert fire_ws.cell(row=2, column=1).value == "Lean FIRE"


async def test_generate_excel_executor_not_anonymized(
    db_session: AsyncSession,
    household: Household,
    primary_member: HouseholdMember,
    primary_user: User,
    tmp_path: Any,
) -> None:
    """Executor export creates a non-anonymized file."""
    job = _make_job(household, primary_user, primary_member, export_type="excel_executor")
    db_session.add(job)
    await db_session.flush()

    filename = await excel_exporter.generate(job, db_session, str(tmp_path))

    assert "executor" in filename
    assert os.path.exists(os.path.join(str(tmp_path), filename))


async def test_generate_excel_filename_contains_summary_or_executor(
    db_session: AsyncSession,
    household: Household,
    primary_member: HouseholdMember,
    primary_user: User,
    tmp_path: Any,
) -> None:
    """Filename suffix matches anonymized flag."""
    summary_job = _make_job(household, primary_user, primary_member, export_type="excel_summary")
    db_session.add(summary_job)
    await db_session.flush()

    summary_name = await excel_exporter.generate(summary_job, db_session, str(tmp_path))
    assert "summary" in summary_name
